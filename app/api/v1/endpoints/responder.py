from io import BytesIO
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.api.v1.dependencies import CurrentUser, require_auth
from app.core.database import get_db
from app.schemas import ResponderCreate, ResponderDetailsResponse, ResponderListItem
from app.services import responder_service

router = APIRouter(prefix="/responders", tags=["responders"])


@router.get("/all", dependencies=[Depends(require_auth)], response_model=list[ResponderListItem])
async def get_all_responders(db:AsyncSession = Depends(get_db)) -> list[ResponderListItem]:
    return await responder_service.get_all_responders(db=db)


@router.get("/additional-details/{responder_id}", dependencies=[Depends(require_auth)], response_model=ResponderDetailsResponse)
async def get_responder_details(responder_id: str, db:AsyncSession = Depends(get_db)) -> ResponderDetailsResponse:
    return await responder_service.get_responder_details(responder_id=responder_id, db=db)


@router.get("/template", dependencies=[Depends(require_auth)])
async def download_responder_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Responders"

    headers = ["First Name", "Last Name", "Phone Number"]
    widths = [20, 20, 25]

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, (header, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 25
    ws.sheet_view.pane = None
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=responder_template.xlsx"},
    )


@router.post("/bulk", response_model=list[ResponderListItem])
async def bulk_create_responders(
    responders: list[ResponderCreate], 
    db: AsyncSession = Depends(get_db),
    user: CurrentUser = Depends(require_auth)) -> list[ResponderListItem]:
    
    return await responder_service.bulk_create_responders(responders=responders, db=db, user_id=user.id)
